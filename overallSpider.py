import requests
import traceback
from lxml import etree
import json
import openpyxl
import time
import pymysql
import mysqlConfig


def GetRowData() -> dict:
    url = "https://voice.baidu.com/act/newpneumonia/newpneumonia#tab"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"}
    response = requests.get(url, headers=headers)
    html = etree.HTML(response.text)
    result = html.xpath("//*[@id=\"captain-config\"]/text()")
    return json.loads(result[0])["component"][0]


def DealTime(intTime: str) -> str:
    if len(intTime) == 10:
        tupTime = time.localtime(int(intTime))
    elif len(intTime) == 13:
        tupTime = time.localtime(float(int(intTime) / 1000))
    stadardTime = time.strftime("%Y-%m-%d %H:%M:%S", tupTime)
    return stadardTime


def GetSumDomData(result: dict) -> list:
    result = result["summaryDataIn"]
    sumDomData = {}
    sumDomKey = {
        "confirmed": "confirmed",
        "died": "died",
        "cured": "cured",
        "asymptomatic": "asymptomatic",
        "asymptomaticRelative": "asymptomatic_relative",
        "unconfirmed": "unconfirmed",
        "relativeTime": "relative_time",
        "confirmedRelative": "confirmed_relative",
        "unconfirmedRelative": "unconfirmed_relative",
        "curedRelative": "cured_relative",
        "diedRelative": "died_relative",
        "icu": "serious",
        "icuRelative": "serious_relative",
        "overseasInput": "foreign_input",
        "unOverseasInputCumulative": "118604",
        "overseasInputRelative": "foreign_input_relative",
        "unOverseasInputNewAdd": "91",
        "curConfirm": "cur_confirm",
        "curConfirmRelative": "cur_confirm_relative",
        "icuDisable": "1"
    }
    for key in result:
        if key in sumDomKey and not sumDomKey[key].isdigit():
            if key == "relativeTime":
                sumDomData[sumDomKey[key]] = DealTime(result[key])
            else:
                sumDomData[sumDomKey[key]] = result[key]
        elif key in sumDomKey and not sumDomKey[key].isdigit():
            sumDomData[key] = result[key]
    temp = []
    temp.append(sumDomData)
    return temp


def GetProvinceData(result: dict) -> list:
    temp = result["caseList"]
    provinceKey = {
        "confirmed": "confirmed",
        "died": "died",
        "crued": "cured",
        "relativeTime": "relative_time",
        "confirmedRelative": "confirm_relative",
        "diedRelative": "died_relative",
        "curedRelative": "cured_relative",
        "asymptomaticRelative": "asymptomatic_relative",
        "asymptomatic": "asymptomatic",
        "nativeRelative": "0",
        "curConfirm": "cur_confirm",
        "curConfirmRelative": "cur_confirm_relative",
        "overseasInputRelative": "0",
        "icuDisable": "1",
        "area": "province"
    }
    for item in temp:
        item.pop("subList")
    provinceData = []
    for i in range(len(temp)):
        provinceData.append({})
        for key in temp[i]:
            if key in provinceKey and not provinceKey[key].isdigit():
                provinceData[i][provinceKey[key]] = DealTime(temp[i][key]) if key == "relativeTime" else (
                    temp[i][key] if temp[i][key] != '' else '0')
            elif key in provinceKey and not provinceKey[key].isdigit():
                provinceData[i][key] = temp[i][key] if temp[i][key] != '' else '0'
    return provinceData


def GetNews() -> list:
    url = "https://sa.sogou.com/new-weball/page/sgs/epidemic?type_page=VR"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"}
    response = requests.get(url, headers=headers)
    response = response.text.encode("ISO-8859-1").decode(requests.utils.get_encodings_from_content(response.text)[0])
    html = etree.HTML(response)
    target = html.xpath("/html/body/script[1]/text()")
    rawNewsData: str = "}," + target[0].split("timeline")[2].split("columns")[0].split("[")[1].split("]")[0] + ",{"
    rawNewsData: list = rawNewsData.split("},{")
    rawNewsData.pop(0)
    rawNewsData.pop()
    for i in range(len(rawNewsData)):
        rawNewsData[i] = "{" + rawNewsData[i]
        rawNewsData[i] += "}"
        rawNewsData[i] = json.loads(rawNewsData[i])
    newsKey = {"timestamp": "time", "url": "link", "title": "title", "content": "content", "source": "source"}
    newsData = []
    for i in range(len(rawNewsData)):
        newsData.append({})
        for key in rawNewsData[i]:
            if key in newsKey and not newsKey[key].isdigit():
                newsData[i][newsKey[key]] = DealTime(str(rawNewsData[i][key])) if key == "timestamp" else \
                    rawNewsData[i][key].strip()
            elif key in newsKey and not newsKey[key].isdigit():
                newsData[i][key] = rawNewsData[i][key].strip()
    return newsData


def SaveResult(data: list, method: str, fileName: str = "", sheetName: str = "") -> None:
    if method == "Console":
        for item in data:
            for key in item:
                print(key + "：" + item[key])
            print()
        print()
    elif method == "Excel":
        if ".xlsx" not in fileName:
            fileName += ".xlsx"
        try:
            file = openpyxl.load_workbook(fileName)
        except:
            file = openpyxl.Workbook()
            file.remove(file["Sheet"])
        if sheetName in file.sheetnames:
            file.remove(file[sheetName])
        sheet = file.create_sheet(sheetName)
        sheet.append(list(data[0].keys()))
        for item in data:
            sheet.append(list(item.values()))
        file.save(fileName)
    elif method == "Text":
        if ".txt" not in fileName:
            fileName += ".txt"
        with open(fileName, "w+", encoding="utf-8") as file_object:
            for item in data:
                for key in item:
                    file_object.write(key + "：" + item[key] + "\n")
                file_object.write("\n")
            file_object.write("\n")
    elif method == "MySQL":
        connection = pymysql.connect(**mysqlConfig.mysql_config)
        cursor = connection.cursor()
        try:
            sql = "CREATE DATABASE IF NOT EXISTS covid;"
            cursor.execute(sql)
            cursor.close()
            connection.close()
            connection = pymysql.connect(**mysqlConfig.mysql_config)
            cursor = connection.cursor()
            sql = """DROP TABLE IF EXISTS """ + fileName + """;"""
            cursor.execute(sql)
            sql = "CREATE TABLE " + fileName + " (id INT PRIMARY KEY,"
            count = 1
            for key in data[0]:
                if key == "relative_time" or key == "province" or key == "time":
                    sql += key + " CHAR(20)"
                elif key == "title":
                    sql += key + " CHAR(100)"
                elif key == "content" or key == "link" or key == "source":
                    sql += key + " TEXT"
                else:
                    sql += key + " INT"
                if count == len(data[0]):
                    sql += ")"
                else:
                    sql += ","
                count += 1
            sql += """ENGINE = InnoDB DEFAULT CHARSET = utf8mb4;"""
            cursor.execute(sql)
            for i in range(len(data)):
                count = 1
                sql = "INSERT INTO " + fileName + " VALUES (" + str(i + 1) + ","
                for key in data[i]:
                    if key == "province":
                        if data[i][key] == "上海":
                            save_Shanghai_result(data[i])
                    if key == "relative_time" or key == "province" or key == "time" or key == "title" or key == "content" or key == "link" or key == "source":
                        sql += "\"" + data[i][key] + "\""
                    else:
                        sql += data[i][key]
                    if count == len(data[i]):
                        sql += ");"
                    else:
                        sql += ","
                    count += 1
                # print(sql)
                cursor.execute(sql)
                connection.commit()
            cursor.close()
            connection.close()
        except Exception as e:
            traceback.print_exc()


def db_create_SH_covid_data():
    db = pymysql.connect(**mysqlConfig.mysql_config)
    cursor = db.cursor()
    create_table = """
                        CREATE TABLE IF NOT EXISTS SH_covid_data (
                        id int NOT NULL AUTO_INCREMENT,
                        confirmed int DEFAULT NULL,
                        asymptomatic int DEFAULT NULL,
                        time varchar(30) DEFAULT NULL,
                        PRIMARY KEY (`id`)
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                        """
    cursor.execute(create_table)
    db.commit()
    cursor.close()
    db.close()


def save_Shanghai_result(obj):
    db = pymysql.connect(**mysqlConfig.mysql_config)
    cursor = db.cursor()
    time = obj['relative_time'].split(' ')[0].split('-', 1)[1]
    covid_data = (obj['confirm_relative'], obj['asymptomatic_relative'], time)
    get_info = """select * from SH_covid_data order by id desc limit 1"""
    cursor.execute(get_info)
    last_day = cursor.fetchone()
    # print(last_day)
    if last_day[3] != time:
        insert_sql = 'INSERT INTO SH_covid_data(confirmed, asymptomatic, time) VALUES(%s, %s, %s)'
        insert_res = cursor.executemany(insert_sql, tuple([(covid_data)]))
        print(insert_res)
    db.commit()
    cursor.close()
    db.close()


if __name__ == '__main__':
    db_create_SH_covid_data()
    rowResult = GetRowData()
    sumDomData = GetSumDomData(rowResult)
    SaveResult(sumDomData, "MySQL", "sumdom")
    provinceData = GetProvinceData(rowResult)
    SaveResult(provinceData, "MySQL", "province")

# alter table SH_covid_data  AUTO_INCREMENT=8;
